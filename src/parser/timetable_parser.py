"""
timetable_parser.py
Reads the institutional timetable Excel file and returns:
  - A list of Room objects
  - A list of TimeSlot objects (ordered, with adjacency indices set)
  - A list of SlotBlock objects (pre-computed consecutive lab blocks)

Expected Excel structure (two possible layouts supported):

Layout A — Flat table (one row per room-slot combination):
    Day | Start Time | End Time | Room | Capacity | Room Type (optional)

Layout B — Grid (institutional timetable style):
    Rooms as rows, time slots as columns (or vice versa), with days on separate sheets.

The parser auto-detects the layout by inspecting the first few rows/columns.

If neither layout is detected, it falls back to a default GIK slot configuration
so development can continue without the actual institutional timetable file.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from src.models.room import Room, RoomType
from src.models.slot import SlotBlock, TimeSlot

# ── Default GIK timetable (fallback when file not provided) ───────────────────
# Derived from the standard GIK scheduling structure (Mon–Fri, 8:00–17:00)

_DEFAULT_DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# Mon-Thu use the standard slot grid (verified against PDF top header row).
_DEFAULT_SLOTS = [
    ("08:00", "08:50"),
    ("09:00", "09:50"),
    # --- Breakfast break 09:50-10:30 ---
    ("10:30", "11:20"),
    ("11:30", "12:20"),
    ("12:30", "13:20"),
    # --- Lunch/Zuhr break 13:20-14:30 ---
    ("14:30", "15:20"),
    ("15:30", "16:20"),
    ("16:30", "17:20"),
]

# Friday morning slots are shifted earlier — the PDF (page 5) explicitly
# shows 10:00-10:50, 11:00-11:50, 12:00-12:50 instead of the 10:30/11:30/12:30
# pattern used Mon-Thu. This compresses morning to fit Jumu'ah prayers.
# Afternoon slots match Mon-Thu.
_FRIDAY_SLOTS = [
    ("08:00", "08:50"),
    ("09:00", "09:50"),
    ("10:00", "10:50"),
    ("11:00", "11:50"),
    ("12:00", "12:50"),
    # --- Jumu'ah break 12:50-14:30 ---
    ("14:30", "15:20"),
    ("15:30", "16:20"),
    ("16:30", "17:20"),
]

# Synced with data/gikirooms.xlsx (authoritative GIK room database).
# Note: gikirooms.xlsx mistakenly spells the FMCE building as "FCME"
# (Faculty of Materials & Chemical Engineering = FMCE). All FCME→FMCE
# normalization is applied here so the rest of the codebase uses one spelling.
_DEFAULT_ROOMS = [
    # FCSE — gikirooms.xlsx is incomplete here; FCSE also has lab rooms
    # (SE Lab, CyS Lab) shared across CS/SE/CY/CE departments per institutional
    # convention. Capacities are estimates pending confirmation from authoritative
    # source. Confirmed by user (CS378, GIK student) on 2026-05-03.
    ("CS LH1",          80, RoomType.LECTURE),
    ("CS LH2",          80, RoomType.LECTURE),
    ("CS LH3",          80, RoomType.LECTURE),
    ("CS LH4",          80, RoomType.LECTURE),
    # FCSE has only SE Lab. CyS / AI / DA labs all live in the Academic Block
    # (see Acad. Block section below). Verified with GIK student.
    ("FCSE - SE Lab",   60, RoomType.LAB),
    # FEE
    ("EE LH1",          60, RoomType.LECTURE),
    ("EE LH2",          80, RoomType.LECTURE),
    ("EE LH3",          80, RoomType.LECTURE),
    ("EE LH4",          80, RoomType.LECTURE),
    ("EE Main",        150, RoomType.LECTURE),
    ("FEE Quiz Hall",  120, RoomType.LECTURE),
    # FES (Faculty of Engineering Sciences — the "FBS" group in the old reference grid)
    ("ES LH1",          60, RoomType.LECTURE),
    ("ES LH2",          60, RoomType.LECTURE),
    ("ES LH3",          60, RoomType.LECTURE),
    ("ES LH4",          60, RoomType.LECTURE),
    ("ES Main",        150, RoomType.LECTURE),
    ("FES Quiz Hall",  120, RoomType.LECTURE),
    ("FES - PH Lab",   100, RoomType.LAB),
    ("FES - PH Lab 2", 100, RoomType.LAB),
    ("FES - SE Lab",    60, RoomType.LAB),
    # Academic Block
    ("AcB LH1",         60, RoomType.LECTURE),
    ("AcB LH2",         60, RoomType.LECTURE),
    ("AcB LH3",         60, RoomType.LECTURE),
    ("AcB LH4",         80, RoomType.LECTURE),
    ("AcB LH5",         80, RoomType.LECTURE),
    ("AcB LH6",         80, RoomType.LECTURE),
    ("AcB LH7",         60, RoomType.LECTURE),
    ("AcB LH8",         80, RoomType.LECTURE),
    ("AcB LH9",         60, RoomType.LECTURE),
    ("AcB LH10",        80, RoomType.LECTURE),
    ("AcB LH11",        60, RoomType.LECTURE),
    ("AcB LH12",        60, RoomType.LECTURE),
    # AcB Main rooms — large lecture halls used for big freshman sections
    # (CS101, CS112 multi-section, CH101, CH161, CV231, MT102 cross-listed).
    ("AcB Main1",      200, RoomType.LECTURE),
    ("AcB Main2",      200, RoomType.LECTURE),
    ("AcB Main3",      200, RoomType.LECTURE),
    ("ACB - AI Lab",    50, RoomType.LAB),
    ("ACB - CYS Lab",   40, RoomType.LAB),
    ("ACB - DA Lab",    50, RoomType.LAB),
    # Sim. Lab is referenced in PDF AcB area for AI/DS simulations
    ("Sim. Lab",        40, RoomType.LAB),
    # Business Block — PDF labels these specific rooms
    ("BB LH2",          80, RoomType.LECTURE),
    ("BB EH1",          80, RoomType.LECTURE),
    ("BB EH2",          80, RoomType.LECTURE),
    ("BB EH3",          80, RoomType.LECTURE),
    ("BB EH4",          80, RoomType.LECTURE),
    ("BB Main",        150, RoomType.LECTURE),
    # PDF shows two distinct PC labs in the BB area, plus Sem. Hall and
    # writing rooms (WR 1, WR 2) used by SC/AF/EM/MGS courses.
    ("Old PC Lab",      40, RoomType.LAB),
    ("New PC Lab",      40, RoomType.LAB),
    ("Sem. Hall",       60, RoomType.LAB),
    ("WR 1",            40, RoomType.LAB),
    ("WR 2",            40, RoomType.LAB),
    # FME
    ("ME LH1",          80, RoomType.LECTURE),
    ("ME LH2",          80, RoomType.LECTURE),
    ("ME LH3",          80, RoomType.LECTURE),
    ("ME Main",        150, RoomType.LECTURE),
    ("FME Quiz Hall",  120, RoomType.LECTURE),
    ("FME Lab",        130, RoomType.LAB),
    # FMCE (file mis-spells as FCME — normalized here)
    ("MCE LH1",         60, RoomType.LECTURE),
    ("MCE LH2",         60, RoomType.LECTURE),
    ("MCE LH3",         60, RoomType.LECTURE),
    ("MCE LH4",         60, RoomType.LECTURE),
    ("MCE Main",       150, RoomType.LECTURE),
    ("FMCE Quiz Hall", 120, RoomType.LECTURE),
    ("FMCE - MM Lab",   40, RoomType.LAB),
    ("FMCE - CH Lab",   40, RoomType.LAB),
    # Mat. Lab — Materials Engineering hands-on lab (PDF labels it under FMCE)
    ("Mat. Lab",        40, RoomType.LAB),
    # FES PC Lab — PDF labels "PC Lab." in the FES section
    ("PC Lab",          30, RoomType.LAB),
    # FBS (cross-faculty lab pool)
    ("FBS Lab",        120, RoomType.LAB),
    # Placeholder for unmapped labs
    ("TBA",            999, RoomType.LAB),
]

# ── Column aliases for Layout A ───────────────────────────────────────────────
_COL_ALIASES: dict[str, list[str]] = {
    "day":       ["Day", "DAY", "Weekday"],
    "start":     ["Start Time", "Start", "START", "From", "Time Start"],
    "end":       ["End Time", "End", "END", "To", "Time End"],
    "room":      ["Room", "ROOM", "Hall", "Classroom", "Venue"],
    "capacity":  ["Capacity", "CAP", "Cap", "Seats", "Strength"],
    "room_type": ["Room Type", "Type", "TYPE", "Category"],
}


def _resolve_columns(df_columns: list[str]) -> dict[str, str | None]:
    col_set = set(df_columns)
    return {
        field: next((a for a in aliases if a in col_set), None)
        for field, aliases in _COL_ALIASES.items()
    }


def _clean_str(val) -> str:
    if pd.isna(val) if not isinstance(val, str) else False:
        return ""
    return str(val).strip()


def _clean_int(val, default: int = 0) -> int:
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return default


def _infer_room_type_from_str(type_str: str) -> RoomType:
    t = type_str.lower()
    if "lab" in t:
        return RoomType.LAB
    if "seminar" in t or "discussion" in t:
        return RoomType.SEMINAR
    return RoomType.LECTURE


class TimetableParser:
    """
    Parses the institutional timetable Excel file into Room and TimeSlot lists.

    Parameters
    ----------
    path : str | Path | None
        Path to the institutional timetable Excel file.
        If None or file doesn't exist, the default GIK configuration is used.
    sheet_name : str | int
        Sheet to read. Defaults to first sheet (0).

    Example
    -------
        parser = TimetableParser("data/institutional_timetable.xlsx")
        rooms  = parser.rooms
        slots  = parser.slots
        blocks = parser.lab_blocks(length=2)
    """

    def __init__(self, path: str | Path | None = None, sheet_name: str | int = 0):
        self.path       = Path(path) if path else None
        self.sheet_name = sheet_name

        self._rooms:  list[Room]     = []
        self._slots:  list[TimeSlot] = []
        self._errors: list[str]      = []

        self._parse()

    # ── Public API ────────────────────────────────────────────────────────────

    @property
    def rooms(self) -> list[Room]:
        """All rooms parsed from the file (or defaults)."""
        return list(self._rooms)

    @property
    def slots(self) -> list[TimeSlot]:
        """All time slots, sorted chronologically."""
        return list(self._slots)

    @property
    def errors(self) -> list[str]:
        """Non-fatal warnings collected during parsing."""
        return list(self._errors)

    def slots_for_day(self, day: str) -> list[TimeSlot]:
        """Return all slots for a specific day, in order."""
        return [s for s in self._slots if s.day == day]

    def rooms_of_type(self, room_type: RoomType) -> list[Room]:
        """Return all rooms of a given RoomType."""
        return [r for r in self._rooms if r.room_type == room_type]

    def lab_blocks(self, length: int = 2) -> list[SlotBlock]:
        """
        Return all valid consecutive slot blocks for lab scheduling.

        A block is valid if 'length' consecutive slots on the same day
        are adjacent (gap ≤ 10 minutes between end of one and start of next).

        Parameters
        ----------
        length : int
            Number of consecutive slots in each lab block. Default is 2.
        """
        blocks: list[SlotBlock] = []
        for day in _DEFAULT_DAYS:
            day_slots = self.slots_for_day(day)
            for i in range(len(day_slots) - length + 1):
                candidate = day_slots[i : i + length]
                # Check all adjacent pairs are consecutive
                if all(
                    candidate[j].is_adjacent_to(candidate[j + 1])
                    for j in range(len(candidate) - 1)
                ):
                    blocks.append(SlotBlock(slots=candidate))
        return blocks

    def summary(self) -> dict:
        """Return a quick summary after parsing."""
        return {
            "rooms":      len(self._rooms),
            "lab_rooms":  len(self.rooms_of_type(RoomType.LAB)),
            "lecture_rooms": len(self.rooms_of_type(RoomType.LECTURE)),
            "slots":      len(self._slots),
            "days":       len({s.day for s in self._slots}),
            "warnings":   len(self._errors),
        }

    # ── Internal parsing ──────────────────────────────────────────────────────

    def _parse(self):
        """Entry point: decide between file parsing and default config."""
        if self.path and self.path.exists():
            try:
                self._parse_file()
                return
            except Exception as exc:
                self._errors.append(
                    f"Failed to parse '{self.path}': {exc}. "
                    f"Falling back to default GIK configuration."
                )
        self._load_defaults()

    def _parse_file(self):
        """Read the Excel file and detect layout automatically."""
        from openpyxl import load_workbook
        wb = load_workbook(self.path, read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))

        # Detect GIK grid layout: first cell of row 0 is a day name
        days = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday"}
        if all_rows and str(all_rows[0][0]).strip() in days:
            self._parse_grid(all_rows)
            return

        # Fall back to flat-table layout
        df = pd.read_excel(self.path, sheet_name=self.sheet_name,
                           header=0, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        col_map = _resolve_columns(list(df.columns))
        if col_map["day"] and col_map["start"] and col_map["room"]:
            self._parse_flat(df, col_map)
        else:
            self._errors.append("Unrecognised timetable layout. Using defaults.")
            self._load_defaults()

    def _parse_grid(self, all_rows: list):
        """
        Parse GIK grid-format timetable.
        Day name is in col 0 of header rows.
        Room names are in col 2 of data rows.
        Time slots are non-None values in col 3+ of header rows.
        """
        import re
        days = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday"}
        time_pattern = re.compile(r"\d{1,2}:\d{2}")

        # Extract rooms from col 2 across all rows
        room_names: list[str] = []
        seen_rooms: set[str] = set()
        for row in all_rows:
            val = row[2] if len(row) > 2 else None
            if val and str(val).strip() and str(val).strip() not in days:
                name = str(val).strip()
                if name not in seen_rooms:
                    seen_rooms.add(name)
                    room_names.append(name)

        # Build rooms — infer type from name
        from src.models.room import Room, RoomType
        rooms_seen: dict[str, Room] = {}
        for name in room_names:
            n_low = name.lower()
            if any(k in n_low for k in ("lab", "pc", "ph lab")):
                rtype = RoomType.LAB
                cap   = 30
            elif "main" in n_low:
                rtype = RoomType.LECTURE
                cap   = 200
            else:
                rtype = RoomType.LECTURE
                cap   = 80
            rooms_seen[name] = Room(name=name, capacity=cap, room_type=rtype)

        # Extract slots per day from header rows
        slots_seen: dict[str, "TimeSlot"] = {}
        from src.models.slot import TimeSlot

        for row in all_rows:
            day_val = str(row[0]).strip() if row[0] else ""
            if day_val not in days:
                continue
            day = day_val
            for cell in row[3:]:
                if not cell:
                    continue
                cell_str = str(cell).strip()
                if not time_pattern.match(cell_str):
                    continue
                # Parse "8:00-8:50" or "10:30-11:20"
                parts = cell_str.replace(" ", "").split("-")
                if len(parts) != 2:
                    continue
                start_raw, end_raw = parts[0], parts[1]
                # Normalise to HH:MM
                def norm(t):
                    h, m = t.split(":")
                    return f"{int(h):02d}:{m}"
                start, end = norm(start_raw), norm(end_raw)
                key = f"{day}_{start}"
                if key not in slots_seen:
                    slots_seen[key] = TimeSlot(day=day, start_time=start, end_time=end)

        self._rooms = list(rooms_seen.values())
        self._slots = self._index_slots(list(slots_seen.values()))

    def _parse_flat(self, df: pd.DataFrame, col_map: dict[str, str | None]):
        """
        Parse a flat-table layout:
            Day | Start Time | End Time | Room | Capacity | Room Type
        """
        rooms_seen:  dict[str, Room]     = {}
        slots_seen:  dict[str, TimeSlot] = {}

        for row_idx, row in df.iterrows():
            day   = _clean_str(row.get(col_map["day"],   ""))
            start = _clean_str(row.get(col_map["start"], ""))
            room  = _clean_str(row.get(col_map["room"],  ""))

            if not day or not start or not room:
                continue

            # Normalise day capitalisation
            day = day.capitalize()
            if day not in _DEFAULT_DAYS:
                self._errors.append(
                    f"Row {row_idx}: unknown day '{day}', skipping."
                )
                continue

            # End time (optional — default to start + 50 min)
            end_col = col_map.get("end")
            end     = _clean_str(row.get(end_col, "")) if end_col else ""
            if not end:
                h, m = int(start[:2]), int(start[3:5])
                m   += 50
                h   += m // 60
                m   %= 60
                end  = f"{h:02d}:{m:02d}"

            # Room
            if room not in rooms_seen:
                cap_col  = col_map.get("capacity")
                capacity = _clean_int(row.get(cap_col, 50)) if cap_col else 50

                type_col  = col_map.get("room_type")
                type_str  = _clean_str(row.get(type_col, "")) if type_col else ""
                room_type = (
                    _infer_room_type_from_str(type_str)
                    if type_str
                    else RoomType.LECTURE   # Room.__post_init__ will re-infer from name
                )
                rooms_seen[room] = Room(
                    name=room, capacity=capacity, room_type=room_type
                )

            # TimeSlot
            slot_key = f"{day}_{start}"
            if slot_key not in slots_seen:
                slots_seen[slot_key] = TimeSlot(
                    day=day, start_time=start, end_time=end
                )

        # Assign chronological indices within each day
        self._rooms = list(rooms_seen.values())
        self._slots = self._index_slots(list(slots_seen.values()))

    def _load_defaults(self):
        """Build rooms and slots from the hardcoded GIK defaults."""
        self._rooms = [
            Room(name=name, capacity=cap, room_type=rtype)
            for name, cap, rtype in _DEFAULT_ROOMS
        ]

        raw_slots: list[TimeSlot] = []
        for day in _DEFAULT_DAYS:
            day_slots = _FRIDAY_SLOTS if day == "Friday" else _DEFAULT_SLOTS
            for start, end in day_slots:
                raw_slots.append(TimeSlot(day=day, start_time=start, end_time=end))

        self._slots = self._index_slots(raw_slots)

    @staticmethod
    def _index_slots(slots: list[TimeSlot]) -> list[TimeSlot]:
        """
        Sort slots chronologically and assign a per-day index to each.
        The index is used for O(1) adjacency lookups in the scheduler.
        """
        sorted_slots = sorted(slots)   # uses TimeSlot.__lt__ via _sort_key

        day_counter: dict[str, int] = {}
        for slot in sorted_slots:
            idx = day_counter.get(slot.day, 0)
            slot.index = idx
            day_counter[slot.day] = idx + 1

        return sorted_slots