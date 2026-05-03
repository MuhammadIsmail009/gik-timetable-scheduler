"""
excel_exporter.py
Exports the generated timetable to Excel matching the GIK reference format.

Output workbook has two sheets:
  1. "Timetable"   — Reference-format grid: 5 day sections × 46 rooms × 8 slots
  2. "Lab Schedule"— Clean table of all lab assignments (virtual rooms not in grid)

Grid layout per day (48 rows):
  Row 0 : Day header + time slot labels
  Rows 1-46 : One row per room (room name in col 2, building in col 1 for first room)
  Row 47 : Blank separator between days
  Cols : 0=Day | 1=Building | 2=Room | 3=8:00 | 4=9:00 | 5=BREAK |
         6=10:30 | 7=11:30 | 8=12:30 | 9=BREAK | 10=14:30 | 11=15:30 | 12=16:30 |
         13=Room(legend) | 14=Building(legend) | 15=Day(legend)

Usage:
    from src.exporter.excel_exporter import ExcelExporter
    ExcelExporter(assignments).save("output/timetable.xlsx")
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# Mon–Thu slot-label → column index (0-based, matches PDF top header row).
SLOT_COL: dict[str, int] = {
    "08:00–08:50": 3,
    "09:00–09:50": 4,
    "10:30–11:20": 6,
    "11:30–12:20": 7,
    "12:30–13:20": 8,
    "14:30–15:20": 10,
    "15:30–16:20": 11,
    "16:30–17:20": 12,
}

# Friday's morning grid is shifted earlier (10:00 / 11:00 / 12:00) per PDF p.5.
# A separate column map is used so cells render on the correct Friday columns.
FRIDAY_SLOT_COL: dict[str, int] = {
    "08:00–08:50": 3,
    "09:00–09:50": 4,
    "10:00–10:50": 5,
    "11:00–11:50": 6,
    "12:00–12:50": 7,
    "14:30–15:20": 10,
    "15:30–16:20": 11,
    "16:30–17:20": 12,
}

# Mon–Thu time slot column headers
SLOT_DISPLAY: dict[int, str] = {
    3:  "8:00-8:50",
    4:  "9:00-9:50",
    6:  "10:30-11:20",
    7:  "11:30-12:20",
    8:  "12:30-13:20",
    10: "14:30-15:20",
    11: "15:30-16:20",
    12: "16:30-17:20",
}

# Friday-specific column headers (5 contiguous morning slots, no breakfast break).
FRIDAY_SLOT_DISPLAY: dict[int, str] = {
    3:  "8:00-8:50",
    4:  "9:00-9:50",
    5:  "10:00-10:50",
    6:  "11:00-11:50",
    7:  "12:00-12:50",
    10: "14:30-15:20",
    11: "15:30-16:20",
    12: "16:30-17:20",
}

# Building groups — order, naming, and room list verified against the
# Spring 2026 PDF (data/TimeTable Spring 2026 *.pdf). Rooms grouped by the
# header label that appears in the PDF's building-group strip.
BUILDING_GROUPS: list[tuple[str, list[str]]] = [
    ("FCSE/FEE", [
        "CS LH1", "CS LH2", "CS LH3", "FCSE - SE Lab",
        "EE LH4", "EE LH5", "EE LH6", "EE Main",
    ]),
    ("FES", [
        "ES LH1", "ES LH2", "ES LH3", "ES Main",
        "FES - PH Lab", "PC Lab",
    ]),
    ("Acad. Block", [
        "AcB LH1",  "AcB LH2",  "AcB LH3",  "AcB LH4",
        "AcB LH5",  "AcB LH6",  "AcB LH7",  "AcB LH8",
        "AcB LH9",  "AcB LH10", "AcB LH11", "AcB LH12",
        "AcB Main1","AcB Main2","AcB Main3",
    ]),
    ("BB", [
        "BB Main", "BB LH2", "BB EH1", "BB EH2", "BB EH3", "BB EH4",
        "Old PC Lab", "New PC Lab", "Sem. Hall", "WR 1", "WR 2",
    ]),
    ("FME", [
        "ME LH1", "ME LH2", "ME LH3", "ME Main", "TBA",
    ]),
    ("FMCE", [
        "MCE LH1", "MCE LH2", "MCE LH3", "MCE LH4", "MCE Main",
        "Mat. Lab",
    ]),
]

# Labs section is built dynamically from actual assignments — see _build_lab_rooms()

# Flat ordered room list (matches reference file row order)
ALL_ROOMS: list[str] = [r for _, rooms in BUILDING_GROUPS for r in rooms]

# Room → building lookup
ROOM_TO_BUILDING: dict[str, str] = {
    room: bldg
    for bldg, rooms in BUILDING_GROUPS
    for room in rooms
}

# Total columns in grid (0..12 — right-side legend removed)
TOTAL_COLS = 13

# Column widths
COL_WIDTHS: dict[int, float] = {
    0: 12,   # Day
    1: 12,   # Building
    2: 12,   # Room
    3: 14,   # 8:00
    4: 14,   # 9:00
    5: 2,    # Break
    6: 14,   # 10:30
    7: 14,   # 11:30
    8: 14,   # 12:30
    9: 2,    # Break
    10: 14,  # 14:30
    11: 14,  # 15:30
    12: 14,  # 16:30
}

# ── Colors ────────────────────────────────────────────────────────────────────
C_DAY_BG     = "1F3864"   # dark navy — day header background
C_DAY_FG     = "FFFFFF"   # white — day header text
C_BLDG_BG    = "2E4D7B"   # medium navy — building label
C_BLDG_FG    = "FFFFFF"
C_ROOM_BG    = "D6E4F7"   # light blue — room name cell
C_ROOM_FG    = "1F3864"
C_SLOT_BG    = "1F3864"   # same as day — time slot header
C_SLOT_FG    = "FFFFFF"
C_BREAK_BG   = "BFD0E8"   # grey-blue — break columns
C_EMPTY_BG   = "F5F8FC"   # near-white — empty cells
C_LEGEND_BG  = "E8F0FA"   # very light blue — right legend
C_LEGEND_FG  = "2E4D7B"

# Department prefix → fill color for course cells
DEPT_FILLS: dict[str, str] = {
    "CS": "DAEEF3", "CE": "D5E8D4", "EE": "FFF2CC",
    "ME": "E8D5F5", "CV": "FFE6E6", "AI": "D5F0FB",
    "DS": "E2F0D9", "CH": "FDEBD0", "HM": "EBF5FB",
    "MS": "F5EEF8", "MM": "E8F8F5", "PH": "FEFBD8",
    "MT": "FDEEF4", "ES": "EAF2FF", "SE": "D5E8D4",
    "CY": "D5E8D4", "IF": "F0F0F0", "SC": "F5EEF8",
    "AF": "F5EEF8", "EL": "FEFBD8",
    "_":  "FFFFFF",
}

DEPT_FONT_COLORS: dict[str, str] = {
    "CS": "0D5E73", "CE": "1B5E20", "EE": "7B6000",
    "ME": "4A148C", "CV": "C62828", "AI": "0D47A1",
    "DS": "33691E", "CH": "E65100", "HM": "1565C0",
    "MS": "6A1B9A", "MM": "00695C", "PH": "827717",
    "MT": "880E4F", "ES": "1A237E", "SE": "1B5E20",
    "CY": "1B5E20", "IF": "424242", "SC": "6A1B9A",
    "AF": "6A1B9A", "_":  "212121",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color: str = "212121", bold: bool = False, size: int = 9) -> Font:
    return Font(name="Calibri", color=hex_color, bold=bold, size=size)


def _border(color: str = "B0C4DE", style: str = "thin") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _center(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _dept_fill(course_code: str) -> tuple[str, str]:
    """Return (bg_hex, fg_hex) for a course code prefix."""
    prefix = course_code[:2].upper()
    return (
        DEPT_FILLS.get(prefix,      DEPT_FILLS["_"]),
        DEPT_FONT_COLORS.get(prefix, DEPT_FONT_COLORS["_"]),
    )


def _cell_text(item: dict) -> str:
    """Format a course assignment as the cell text shown in the grid."""
    code    = item.get("course", "")
    section = item.get("section", "")
    # e.g. "CS221 C" or "EE201/EE213"
    return f"{code} {section}".strip()


def _is_virtual_room(room_name: str) -> bool:
    """Return True if this room is a virtual lab room not in the physical grid."""
    return room_name not in ALL_ROOMS and room_name != "TBA"


def _is_numbered_lab(room_name: str) -> tuple[str, int] | None:
    """
    If room_name is a numbered virtual lab instance, return (base, number).
    Recognises both "FCSE - SE Lab 3" and "FCSE - SE Lab #3" forms.
    """
    if not room_name or room_name == "TBA":
        return None
    # "Foo #3" → ("Foo", 3)
    if " #" in room_name:
        base, num = room_name.rsplit(" #", 1)
        if num.isdigit():
            return base, int(num)
    # "Foo 3" → ("Foo", 3) — only if not the bare physical room name
    parts = room_name.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].isdigit() and room_name not in ALL_ROOMS:
        return parts[0], int(parts[1])
    return None


# ── Main class ────────────────────────────────────────────────────────────────

class ExcelExporter:
    """
    Converts scheduler result into a formatted Excel workbook.

    Parameters
    ----------
    assignments : list[dict]
        Each dict = one scheduled course. Expected keys:
        course, section, title, instructor, is_lab, day, slot, room,
        for_batch, students  (matches Course.to_dict())
    """

    def __init__(self, assignments: list[dict]):
        self._data = assignments
        self._wb   = Workbook()

        # Build lookup: (day, slot_label, room_name) → item
        # Each numbered virtual lab instance (e.g. "FCSE - SE Lab 1",
        # "FCSE - SE Lab 2") gets its OWN row in the grid, so labs
        # never appear to overlap visually.
        self._grid: dict[tuple, list[dict]] = {}
        for item in assignments:
            room_name = item.get("room", "")
            key = (item.get("day", ""), item.get("slot", ""), room_name)
            self._grid.setdefault(key, []).append(item)

        # Build the dynamic "Labs" section based on which virtual rooms
        # actually got used in this schedule.
        self._lab_rooms = self._build_lab_rooms()

    def _build_lab_rooms(self) -> list[str]:
        """
        Scan all assignments and collect unique virtual lab rooms — those
        not in the static BUILDING_GROUPS grid. Returns the actual room
        names as they appear in assignments (e.g. "FCSE - SE Lab #2",
        "CV Lab 1") so that `(day, slot, room)` lookups work directly.
        """
        seen: set[str] = set()
        for item in self._data:
            if not item.get("is_lab"):
                continue
            room = item.get("room", "")
            if not room or room in ALL_ROOMS or room == "TBA":
                continue
            seen.add(room)

        # Sort by base name then instance number for stable output
        def sort_key(name: str) -> tuple:
            parsed = _is_numbered_lab(name)
            if parsed is None:
                return (name, 0)
            base, num = parsed
            return (base, num)

        return sorted(seen, key=sort_key)

    # ── Public ────────────────────────────────────────────────────────────────

    def save(self, path: str | Path = "output/timetable.xlsx") -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        self._build_timetable_sheet()

        self._wb.save(path)
        return path

    # ── Sheet 1: Timetable Grid ───────────────────────────────────────────────

    def _build_timetable_sheet(self):
        ws            = self._wb.active
        ws.title      = "Timetable"
        ws.sheet_view.showGridLines = False

        # Set column widths
        for col_idx, width in COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

        current_row = 1   # 1-based openpyxl row

        for day in DAYS:
            # ── Day header row ────────────────────────────────────────────────
            self._write_day_header(ws, current_row, day)
            current_row += 1

            # ── Room rows ─────────────────────────────────────────────────────
            for bldg, rooms in BUILDING_GROUPS:
                for room_idx, room in enumerate(rooms):
                    self._write_room_row(
                        ws, current_row, day, room,
                        bldg if room_idx == 0 else None,
                    )
                    current_row += 1

            # ── Labs section (dynamically built) ──────────────────────────────
            for room_idx, room in enumerate(self._lab_rooms):
                self._write_room_row(
                    ws, current_row, day, room,
                    "Labs" if room_idx == 0 else None,
                )
                current_row += 1

            # ── Blank separator ───────────────────────────────────────────────
            for col in range(TOTAL_COLS):
                cell = ws.cell(row=current_row, column=col + 1)
                cell.fill = _fill("FFFFFF")
            ws.row_dimensions[current_row].height = 6
            current_row += 1

        # Freeze panes at D2 (keep day/room headers fixed)
        ws.freeze_panes = "D2"

    def _write_day_header(self, ws, row: int, day: str):
        """Write the day header row with time slot labels (Friday uses
        the shifted morning grid)."""
        ws.row_dimensions[row].height = 22

        slot_display = FRIDAY_SLOT_DISPLAY if day == "Friday" else SLOT_DISPLAY

        # Col 0: Day name
        c = ws.cell(row=row, column=1, value=day.upper())
        c.fill      = _fill(C_DAY_BG)
        c.font      = _font(C_DAY_FG, bold=True, size=11)
        c.alignment = _center()
        c.border    = _border("FFFFFF")

        # Col 1-2: blank header
        for col in (2, 3):
            c = ws.cell(row=row, column=col)
            c.fill = _fill(C_DAY_BG)

        # Time slot labels — column 5 (breakfast break) is filled on Mon–Thu
        # and used as a real slot column on Friday.
        for col_idx, label in slot_display.items():
            c = ws.cell(row=row, column=col_idx + 1, value=label)
            c.fill      = _fill(C_SLOT_BG)
            c.font      = _font(C_SLOT_FG, bold=True, size=9)
            c.alignment = _center()
            c.border    = _border("FFFFFF")

        # Break columns: only column 9 (lunch) on every day; column 5 is a
        # real slot on Friday (10:00) and a break on Mon–Thu.
        break_cols = (9,) if day == "Friday" else (5, 9)
        for break_col in break_cols:
            c = ws.cell(row=row, column=break_col + 1)
            c.fill = _fill(C_BREAK_BG)

    def _write_room_row(self, ws, row: int, day: str, room: str, building: str | None):
        """Write one room row with all slot cells (Friday uses the shifted
        morning column layout)."""
        ws.row_dimensions[row].height = 18

        slot_col_map = FRIDAY_SLOT_COL if day == "Friday" else SLOT_COL
        break_cols   = (9,) if day == "Friday" else (5, 9)

        # Col 0: Day (blank)
        ws.cell(row=row, column=1).fill = _fill("FFFFFF")

        # Col 1: Building (only for first room in group)
        c = ws.cell(row=row, column=2)
        if building:
            c.value     = building
            c.fill      = _fill(C_BLDG_BG)
            c.font      = _font(C_BLDG_FG, bold=True, size=9)
            c.alignment = _center()
        else:
            c.fill = _fill("FFFFFF")

        # Col 2: Room name
        c = ws.cell(row=row, column=3, value=room)
        c.fill      = _fill(C_ROOM_BG)
        c.font      = _font(C_ROOM_FG, bold=True, size=9)
        c.alignment = _center()
        c.border    = _border()

        # Time slot cells — exact-room match only. Virtual instances
        # ("FCSE - SE Lab #2") are listed separately in the dynamic
        # Labs section so each appears in its own row.
        for slot_label, col_idx in slot_col_map.items():
            items = self._grid.get((day, slot_label, room), [])

            c = ws.cell(row=row, column=col_idx + 1)
            c.border    = _border()
            c.alignment = _center()

            if items:
                text = " / ".join(_cell_text(i) for i in items)
                bg, fg = _dept_fill(items[0].get("course", "_"))
                c.value = text
                c.fill  = _fill(bg)
                c.font  = _font(fg, bold=True, size=9)
            else:
                c.fill = _fill(C_EMPTY_BG)
                c.font = _font(size=9)

        # Break columns
        for break_col in break_cols:
            c = ws.cell(row=row, column=break_col + 1)
            c.fill = _fill(C_BREAK_BG)



    # ── Sheet 2: Lab Schedule ─────────────────────────────────────────────────

    def _build_lab_sheet(self):
        """
        Clean table of all lab assignments.
        Labs assigned to virtual rooms (not in the physical grid) are shown here.
        """
        ws = self._wb.create_sheet("Lab Schedule")
        ws.sheet_view.showGridLines = False

        labs = [d for d in self._data if d.get("is_lab")]

        # Header
        headers = [
            "Course Code", "Section", "Course Title", "Instructor",
            "Day", "Time Slot", "Lab Room", "Batch",
        ]
        col_widths = [13, 9, 36, 28, 12, 18, 28, 12]

        for ci, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
            c = ws.cell(row=1, column=ci, value=hdr)
            c.fill      = _fill(C_DAY_BG)
            c.font      = _font(C_DAY_FG, bold=True, size=10)
            c.alignment = _center()
            c.border    = _border()
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.row_dimensions[1].height = 20
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.freeze_panes = "A2"

        # Sort by day → slot → course
        day_order = {d: i for i, d in enumerate(DAYS)}
        labs_sorted = sorted(
            labs,
            key=lambda d: (
                day_order.get(d.get("day", ""), 9),
                d.get("slot", ""),
                d.get("course", ""),
            ),
        )

        for ri, item in enumerate(labs_sorted, start=2):
            bg, fg = _dept_fill(item.get("course", "_"))

            # Clean up virtual room name (strip instance number)
            room_raw = item.get("room", "")
            # "FCSE - SE Lab 3" → "FCSE - SE Lab"
            room_display = " ".join(
                part for part in room_raw.rsplit(" ", 1)
                if not part.isdigit()
            ) if room_raw.rsplit(" ", 1)[-1].isdigit() else room_raw

            values = [
                item.get("course", ""),
                item.get("section", ""),
                item.get("title", ""),
                item.get("instructor", ""),
                item.get("day", ""),
                item.get("slot", ""),
                room_display,
                item.get("for_batch", ""),
            ]

            for ci, val in enumerate(values, start=1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill      = _fill(bg)
                c.font      = _font(fg, size=9)
                c.border    = _border()
                c.alignment = _left()

            ws.row_dimensions[ri].height = 15

        # Summary at bottom
        total_row = len(labs_sorted) + 3
        ws.cell(row=total_row, column=1, value="Total Lab Sessions").font = _font(C_DAY_BG, bold=True, size=10)
        ws.cell(row=total_row, column=2, value=len(labs_sorted)).font     = _font(C_DAY_BG, bold=True, size=10)